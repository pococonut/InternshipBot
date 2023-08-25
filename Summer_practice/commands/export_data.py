import openpyxl
from create import dp
from aiogram import types
from keyboard import exp_ikb
from db.commands import stud_approve, select_added_users, select_task, select_user, select_all_users


@dp.callback_query_handler(text='export')
async def export(callback: types.CallbackQuery):
    """
    Функция возвращающая клавиатуру со списком данных, доступных для получения в виде excel-файла.
    """
    await callback.message.edit_text("Выберите команду.", parse_mode='HTML', reply_markup=exp_ikb)


@dp.callback_query_handler(text='exp_task')
async def export_task(callback: types.CallbackQuery):
    """
    Функция возвращающая excel-файл с параметрами добавленных задач.
    """
    tasks = select_task()

    if not tasks:
        await callback.message.edit_text("Данные отсутствуют.", parse_mode='HTML', reply_markup=exp_ikb)
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Задачи"

        c2 = sheet.cell(row=1, column=1)
        c2.value = "Сотрудник"
        c3 = sheet.cell(row=1, column=2)
        c3.value = "Студент"
        c4 = sheet.cell(row=1, column=3)
        c4.value = "Название"
        c5 = sheet.cell(row=1, column=4)
        c5.value = "Цель"
        c6 = sheet.cell(row=1, column=5)
        c6.value = "Описание"
        c7 = sheet.cell(row=1, column=6)
        c7.value = "Задачи"
        c8 = sheet.cell(row=1, column=7)
        c8.value = "Требования к технологиям"
        c9 = sheet.cell(row=1, column=8)
        c9.value = "Новые навыки"
        c10 = sheet.cell(row=1, column=9)
        c10.value = "Количество людей"
        c11 = sheet.cell(row=1, column=10)
        c11.value = "Материалы"

        for i in range(2, len(tasks) + 2):
            index = i - 2

            s2 = sheet.cell(row=i, column=1)
            s2.value = select_user(tasks[index].from_id).name if (tasks[index].from_id is not None and select_user(tasks[index].from_id) is not None) else None

            s3 = sheet.cell(row=i, column=2)
            s3.value = select_user(tasks[index].student_id).name if (tasks[index].student_id is not None and select_user(tasks[index].student_id) is not None) else None

            s4 = sheet.cell(row=i, column=3)
            s4.value = tasks[index].task_name

            s5 = sheet.cell(row=i, column=4)
            s5.value = tasks[index].task_goal

            s6 = sheet.cell(row=i, column=5)
            s6.value = tasks[index].task_description

            s7 = sheet.cell(row=i, column=6)
            s7.value = tasks[index].task_tasks

            s8 = sheet.cell(row=i, column=7)
            s8.value = tasks[index].task_technologies

            s9 = sheet.cell(row=i, column=8)
            s9.value = tasks[index].task_new_skills

            s10 = sheet.cell(row=i, column=9)
            s10.value = tasks[index].num_people

            s11 = sheet.cell(row=i, column=10)
            s11.value = tasks[index].materials

        wb.save(r"files\Задачи.xlsx")
        doc = open(r"files\Задачи.xlsx", 'rb')

        await callback.answer()
        await callback.message.answer_document(doc)


@dp.callback_query_handler(text='exp_worker')
async def export_worker(callback: types.CallbackQuery):
    """
    Функция возвращающая excel-файл с авторизированными сотрудниками.
    """
    workers = select_all_users()
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Сотрудники"
    f = 0

    c1 = sheet.cell(row=1, column=1)
    c1.value = "Тип"
    c2 = sheet.cell(row=1, column=2)
    c2.value = "Имя"
    c3 = sheet.cell(row=1, column=3)
    c3.value = "Номер Телефона"
    c4 = sheet.cell(row=1, column=4)
    c4.value = "Дата регистрации"
    c5 = sheet.cell(row=1, column=5)
    c5.value = "Дата изменения"
    c6 = sheet.cell(row=1, column=6)
    c6.value = "Логин"
    c7 = sheet.cell(row=1, column=7)
    c7.value = "Пароль"

    i = 2
    for w in range(2, len(workers) + 2):
        index = w - 2

        if workers[index].type != 'student':
            f = 1
            s1 = sheet.cell(row=i, column=1)
            s1.value = workers[index].type

            s2 = sheet.cell(row=i, column=2)
            s2.value = workers[index].name

            s3 = sheet.cell(row=i, column=3)
            s3.value = workers[index].phone

            s4 = sheet.cell(row=i, column=4)
            s4.value = workers[index].reg_date

            s5 = sheet.cell(row=i, column=5)
            s5.value = workers[index].upd_date

            s6 = sheet.cell(row=i, column=6)
            s6.value = workers[index].login

            s7 = sheet.cell(row=i, column=7)
            s7.value = workers[index].password

            i += 1

    if not f:
        await callback.message.edit_text("Данные отсутствуют.", parse_mode='HTML', reply_markup=exp_ikb)
    else:
        wb.save(r"files\Сотрудники.xlsx")
        doc = open(r"files\Сотрудники.xlsx", 'rb')

    await callback.answer()
    await callback.message.answer_document(doc)


@dp.callback_query_handler(text='exp_appl')
async def export_applications(callback: types.CallbackQuery):
    """
    Функция возвращающая excel-файл с параметрами заявок студентов.
    """
    students = select_all_users()
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Заявки"
    f = 0

    c1 = sheet.cell(row=1, column=1)
    c1.value = "ФИО"
    c2 = sheet.cell(row=1, column=2)
    c2.value = "Номер Телефона"
    c3 = sheet.cell(row=1, column=3)
    c3.value = "ВУЗ"
    c4 = sheet.cell(row=1, column=4)
    c4.value = "Факультет"
    c5 = sheet.cell(row=1, column=5)
    c5.value = "Направление"
    c6 = sheet.cell(row=1, column=6)
    c6.value = "Кафедра"
    c7 = sheet.cell(row=1, column=7)
    c7.value = "Курс"
    c8 = sheet.cell(row=1, column=8)
    c8.value = "Группа"
    c9 = sheet.cell(row=1, column=9)
    c9.value = "Курсовые"
    c10 = sheet.cell(row=1, column=10)
    c10.value = "Знания"

    i = 2
    for a in range(2, len(students) + 2):
        index = a - 2

        if students[index].type == 'student':
            f = 1

            s1 = sheet.cell(row=i, column=1)
            s1.value = students[index].student_name

            s2 = sheet.cell(row=i, column=2)
            s2.value = students[index].phone

            s3 = sheet.cell(row=i, column=3)
            s3.value = students[index].university

            s4 = sheet.cell(row=i, column=4)
            s4.value = students[index].faculty

            s5 = sheet.cell(row=i, column=5)
            s5.value = students[index].specialties

            s6 = sheet.cell(row=i, column=6)
            s6.value = students[index].department

            s7 = sheet.cell(row=i, column=7)
            s7.value = students[index].course

            s8 = sheet.cell(row=i, column=8)
            s8.value = students[index].group

            s9 = sheet.cell(row=i, column=9)
            s9.value = students[index].coursework

            s10 = sheet.cell(row=i, column=10)
            s10.value = students[index].knowledge

            i += 1
    if not f:
        await callback.message.edit_text("Данные отсутствуют.", parse_mode='HTML', reply_markup=exp_ikb)
    else:
        wb.save(r"files\Заявки.xlsx")
        doc = open(r"files\Заявки.xlsx", 'rb')

        await callback.answer()
        await callback.message.answer_document(doc)


@dp.callback_query_handler(text='exp_approved')
async def export_approved(callback: types.CallbackQuery):
    """
    Функция возвращающая excel-файл с параметрами одобренных заявок студентов.
    """
    students = select_all_users()
    if not students:
        await callback.message.edit_text("Данные отсутствуют.", parse_mode='HTML', reply_markup=exp_ikb)
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Заявки"

        c1 = sheet.cell(row=1, column=1)
        c1.value = "ФИО"
        c2 = sheet.cell(row=1, column=2)
        c2.value = "Номер Телефона"
        c3 = sheet.cell(row=1, column=3)
        c3.value = "ВУЗ"
        c4 = sheet.cell(row=1, column=4)
        c4.value = "Факультет"
        c5 = sheet.cell(row=1, column=5)
        c5.value = "Направление"
        c6 = sheet.cell(row=1, column=6)
        c6.value = "Кафедра"
        c7 = sheet.cell(row=1, column=7)
        c7.value = "Курс"
        c8 = sheet.cell(row=1, column=8)
        c8.value = "Группа"
        c9 = sheet.cell(row=1, column=9)
        c9.value = "Курсовые"
        c10 = sheet.cell(row=1, column=10)
        c10.value = "Знания"
        f = 1
        i = 2
        for s in range(2, len(students) + 2):
            index = s - 2

            result = stud_approve(students[index].telegram_id)
            if result is None:
                f = 0
            else:
                if result[0]:
                    s1 = sheet.cell(row=i, column=1)
                    s1.value = students[index].student_name

                    s2 = sheet.cell(row=i, column=2)
                    s2.value = students[index].phone

                    s3 = sheet.cell(row=i, column=3)
                    s3.value = students[index].university

                    s4 = sheet.cell(row=i, column=4)
                    s4.value = students[index].faculty

                    s5 = sheet.cell(row=i, column=5)
                    s5.value = students[index].specialties

                    s6 = sheet.cell(row=i, column=6)
                    s6.value = students[index].department

                    s7 = sheet.cell(row=i, column=7)
                    s7.value = students[index].course

                    s8 = sheet.cell(row=i, column=8)
                    s8.value = students[index].group

                    s9 = sheet.cell(row=i, column=9)
                    s9.value = students[index].coursework

                    s10 = sheet.cell(row=i, column=10)
                    s10.value = students[index].knowledge

                    i += 1
        if f:
            wb.save(r"files\Принятые студенты.xlsx")
            doc = open(r"files\Принятые студенты.xlsx", 'rb')
            await callback.answer()
            await callback.message.answer_document(doc)
        else:
            await callback.message.edit_text("Данные отсутствуют.", parse_mode='HTML', reply_markup=exp_ikb)


@dp.callback_query_handler(text='exp_added')
async def export_added(callback: types.CallbackQuery):
    """
    Функция возвращающая excel-файл с добавленными аккаунтами.
    """
    added_users = select_added_users()
    if not added_users:
        await callback.message.edit_text("Данные отсутствуют.", parse_mode='HTML', reply_markup=exp_ikb)
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Добавленные аккаунты"

        c1 = sheet.cell(row=1, column=1)
        c1.value = "Логин"
        c2 = sheet.cell(row=1, column=2)
        c2.value = "Пароль"
        c3 = sheet.cell(row=1, column=3)
        c3.value = "Тип"
        c4 = sheet.cell(row=1, column=4)
        c4.value = "ФИО"

        for i in range(2, len(added_users) + 2):
            index = i - 2

            s1 = sheet.cell(row=i, column=1)
            s1.value = added_users[index].login

            s2 = sheet.cell(row=i, column=2)
            s2.value = added_users[index].password

            s3 = sheet.cell(row=i, column=3)
            s3.value = added_users[index].type

            s4 = sheet.cell(row=i, column=4)
            s4.value = added_users[index].name_usr

        wb.save(r"files\Добавленные аккаунты.xlsx")
        doc = open(r"files\Добавленные аккаунты.xlsx", 'rb')

        await callback.answer()
        await callback.message.answer_document(doc)
