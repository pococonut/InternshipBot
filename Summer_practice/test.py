# import openpyxl module
import openpyxl

from db.commands import select_task


tasks = select_task()
wb = openpyxl.Workbook()

sheet = wb.active
sheet.title = "Задачи"

c1 = sheet.cell(row=1, column=1)
c1.value = "ID"
c2 = sheet.cell(row=1, column=2)
c2.value = "Сотрудник"
c3 = sheet.cell(row=1, column=3)
c3.value = "Студент"
c4 = sheet.cell(row=1, column=4)
c4.value = "Название"
c5 = sheet.cell(row=1, column=5)
c5.value = "Цель"
c6 = sheet.cell(row=1, column=6)
c6.value = "Описание"
c7 = sheet.cell(row=1, column=7)
c7.value = "Задачи"
c8 = sheet.cell(row=1, column=8)
c8.value = "Требования к технологиям"
c9 = sheet.cell(row=1, column=9)
c9.value = "Новые навыки"
c10 = sheet.cell(row=1, column=10)
c10.value = "Количество людей"
c11 = sheet.cell(row=1, column=11)
c11.value = "Материалы"

print(len(tasks))
for i in range(2, len(tasks)+1):
    index = i-2
    print(tasks[index].student_id)

    s1 = sheet.cell(row=i, column=1)
    s1.value = i-1

    s2 = sheet.cell(row=i, column=2)
    s2.value = tasks[index].from_id

    s3 = sheet.cell(row=i, column=3)
    s3.value = tasks[index].student_id

    s4 = sheet.cell(row=i, column=4)
    s4.value = tasks[index].task_name

    s5 = sheet.cell(row=i, column=5)
    s5.value = tasks[index].task_goal

    s6 = sheet.cell(row=i, column=6)
    s6.value = tasks[index].task_description

    s7 = sheet.cell(row=i, column=7)
    s7.value = tasks[index].task_tasks

    s8 = sheet.cell(row=i, column=8)
    s8.value = tasks[index].task_technologies

    s9 = sheet.cell(row=i, column=9)
    s9.value = tasks[index].task_new_skills

    s10 = sheet.cell(row=i, column=10)
    s10.value = tasks[index].num_people

    s11 = sheet.cell(row=i, column=11)
    s11.value = tasks[index].materials



wb.save(r"C:\Users\polin\Desktop\demo1.xlsx")