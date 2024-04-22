import datetime

import openpyxl
from aiogram import types

from commands.applications import get_students
from create import dp
from keyboard import exp_ikb
from commands.get_menu import callback_check_authentication
from db.commands import stud_approve, select_added_users, select_task, select_user, select_all_users


@dp.callback_query_handler(text='export')
@callback_check_authentication
async def export(callback: types.CallbackQuery):
    """
    Функция возвращающая клавиатуру со списком данных, доступных для получения в виде excel-файла.
    """

    await callback.message.edit_text("Выберите команду.", reply_markup=exp_ikb)


@dp.callback_query_handler(text='exp_task')
@callback_check_authentication
async def export_task(callback: types.CallbackQuery):
    """
    Функция возвращающая excel-файл с параметрами добавленных задач.
    """

    tasks = select_task()
    if not tasks:
        await callback.message.edit_text("Данные отсутствуют.", reply_markup=exp_ikb)
        return

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Задачи"
    columns = ["Сотрудник", "Студент", "Название", "Цель", "Описание", "Задачи",
               "Требования к технологиям", "Новые навыки", "Количество людей", "Материалы"]

    for i in range(1, len(columns)+1):
        c = sheet.cell(row=1, column=i)
        c.value = columns[i-1]

    last_idx = 0
    for i in range(2, len(tasks) + 2):
        index = i - 2

        s2 = sheet.cell(row=i, column=1)
        worker = select_user(tasks[index].from_id)
        task_from = tasks[index].from_id
        if task_from and worker:
            worker_name = worker.name
        else:
            worker_name = None
        s2.value = worker_name

        s3 = sheet.cell(row=i, column=2)
        if tasks[index].student_id and select_user(tasks[index].student_id):
            s3.value = select_user(tasks[index].student_id).name
        else:
            s3.value = None

        s4 = sheet.cell(row=i, column=3)
        s4.value = tasks[index].task_name

        s5 = sheet.cell(row=i, column=4)
        s5.value = tasks[index].task_goal

        s6 = sheet.cell(row=i, column=5)
        s6.value = tasks[index].task_description

        s7 = sheet.cell(row=i, column=6)
        s7.value = tasks[index].task_tasks

        s8 = sheet.cell(row=i, column=7)
        s8.value = tasks[index].task_technologies

        s9 = sheet.cell(row=i, column=8)
        s9.value = tasks[index].task_new_skills

        s10 = sheet.cell(row=i, column=9)
        s10.value = tasks[index].num_people

        s11 = sheet.cell(row=i, column=10)
        s11.value = tasks[index].materials

        last_idx = i + 2

    s_data = sheet.cell(row=last_idx, column=1)
    s_data.value = "Дата экспорта:"
    s_data = sheet.cell(row=last_idx, column=2)
    s_data.value = datetime.date.today()

    wb.save("files/Задачи.xlsx")
    doc = open("files/Задачи.xlsx", 'rb')

    await callback.answer()
    await callback.message.answer_document(doc)


@dp.callback_query_handler(text='exp_worker')
@callback_check_authentication
async def export_worker(callback: types.CallbackQuery):
    """
    Функция возвращающая excel-файл с авторизированными сотрудниками.
    """
    workers = select_all_users()
    if not workers:
        await callback.message.edit_text("Данные отсутствуют.", reply_markup=exp_ikb)
        return

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Сотрудники"
    columns = ["Тип", "Имя", "Номер Телефона", "Дата регистрации", "Дата изменения", "Логин", "Пароль"]

    for i in range(1, len(columns)+1):
        c = sheet.cell(row=1, column=i)
        c.value = columns[i-1]

    i = 2
    for w in range(2, len(workers) + 2):
        index = w - 2

        if workers[index].type != 'student':
            s1 = sheet.cell(row=i, column=1)
            s1.value = workers[index].type

            s2 = sheet.cell(row=i, column=2)
            s2.value = workers[index].name

            s3 = sheet.cell(row=i, column=3)
            s3.value = workers[index].phone

            s4 = sheet.cell(row=i, column=4)
            s4.value = workers[index].reg_date

            s5 = sheet.cell(row=i, column=5)
            s5.value = workers[index].upd_date

            s6 = sheet.cell(row=i, column=6)
            s6.value = workers[index].login

            s7 = sheet.cell(row=i, column=7)
            s7.value = workers[index].password

            i += 1

    s_data = sheet.cell(row=i + 1, column=1)
    s_data.value = "Дата экспорта:"
    s_data = sheet.cell(row=i+1, column=2)
    s_data.value = datetime.date.today()

    wb.save("files/Сотрудники.xlsx")
    doc = open(r"files/Сотрудники.xlsx", 'rb')

    await callback.answer()
    await callback.message.answer_document(doc)


@dp.callback_query_handler(text='exp_appl')
@callback_check_authentication
async def export_applications(callback: types.CallbackQuery):
    """
    Функция возвращающая excel-файл с параметрами заявок студентов.
    """

    students = get_students()
    if not students:
        await callback.message.edit_text("Данные отсутствуют.", reply_markup=exp_ikb)
        return

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Заявки"
    columns = ["ФИО", "Номер Телефона", "ВУЗ", "Факультет", "Направление",
               "Кафедра", "Курс", "Группа", "Курсовые", "Знания"]

    for i in range(1, len(columns) + 1):
        c = sheet.cell(row=1, column=i)
        c.value = columns[i - 1]

    i = 2
    for a in range(2, len(students) + 2):
        index = a - 2

        if students[index].type == 'student':

            s1 = sheet.cell(row=i, column=1)
            s1.value = students[index].student_name

            s2 = sheet.cell(row=i, column=2)
            s2.value = students[index].phone

            s3 = sheet.cell(row=i, column=3)
            s3.value = students[index].university

            s4 = sheet.cell(row=i, column=4)
            s4.value = students[index].faculty

            s5 = sheet.cell(row=i, column=5)
            s5.value = students[index].specialties

            s6 = sheet.cell(row=i, column=6)
            s6.value = students[index].department

            s7 = sheet.cell(row=i, column=7)
            s7.value = students[index].course

            s8 = sheet.cell(row=i, column=8)
            s8.value = students[index].group

            s9 = sheet.cell(row=i, column=9)
            s9.value = students[index].coursework

            s10 = sheet.cell(row=i, column=10)
            s10.value = students[index].knowledge

            i += 1

    s_data = sheet.cell(row=i + 1, column=1)
    s_data.value = "Дата экспорта:"
    s_data = sheet.cell(row=i + 1, column=2)
    s_data.value = datetime.date.today()

    wb.save("files/Заявки.xlsx")
    doc = open("files/Заявки.xlsx", 'rb')

    await callback.answer()
    await callback.message.answer_document(doc)


@dp.callback_query_handler(text='exp_approved')
@callback_check_authentication
async def export_approved(callback: types.CallbackQuery):
    """
    Функция возвращающая excel-файл с параметрами одобренных заявок студентов.
    """
    students = select_all_users()
    if not students:
        await callback.message.edit_text("Данные отсутствуют.", parse_mode='HTML', reply_markup=exp_ikb)
        return

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Принятые студенты"
    columns = ["ФИО", "Номер Телефона", "ВУЗ", "Факультет", "Направление",
               "Кафедра", "Курс", "Группа", "Курсовые", "Знания"]

    for i in range(1, len(columns) + 1):
        c = sheet.cell(row=1, column=i)
        c.value = columns[i - 1]

    i = 2
    for s in range(2, len(students) + 2):
        index = s - 2

        result = stud_approve(students[index].telegram_id)

        if result:
            s1 = sheet.cell(row=i, column=1)
            s1.value = students[index].student_name

            s2 = sheet.cell(row=i, column=2)
            s2.value = students[index].phone

            s3 = sheet.cell(row=i, column=3)
            s3.value = students[index].university

            s4 = sheet.cell(row=i, column=4)
            s4.value = students[index].faculty

            s5 = sheet.cell(row=i, column=5)
            s5.value = students[index].specialties

            s6 = sheet.cell(row=i, column=6)
            s6.value = students[index].department

            s7 = sheet.cell(row=i, column=7)
            s7.value = students[index].course

            s8 = sheet.cell(row=i, column=8)
            s8.value = students[index].group

            s9 = sheet.cell(row=i, column=9)
            s9.value = students[index].coursework

            s10 = sheet.cell(row=i, column=10)
            s10.value = students[index].knowledge

            i += 1

    s_data = sheet.cell(row=i + 1, column=1)
    s_data.value = "Дата экспорта:"
    s_data = sheet.cell(row=i + 1, column=2)
    s_data.value = datetime.date.today()

    wb.save("files/Принятые студенты.xlsx")
    doc = open("files/Принятые студенты.xlsx", 'rb')

    await callback.answer()
    await callback.message.answer_document(doc)


@dp.callback_query_handler(text='exp_added')
@callback_check_authentication
async def export_added(callback: types.CallbackQuery):
    """
    Функция возвращающая excel-файл с добавленными аккаунтами.
    """
    added_users = select_added_users()
    if not added_users:
        await callback.message.edit_text("Данные отсутствуют.", parse_mode='HTML', reply_markup=exp_ikb)
        return

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Добавленные аккаунты"
    columns = ["Логин", "Пароль", "Тип", "ФИО"]
    last_idx = 0

    for i in range(1, len(columns) + 1):
        c = sheet.cell(row=1, column=i)
        c.value = columns[i - 1]

    for i in range(2, len(added_users) + 2):
        index = i - 2

        s1 = sheet.cell(row=i, column=1)
        s1.value = added_users[index].login

        s2 = sheet.cell(row=i, column=2)
        s2.value = added_users[index].password

        s3 = sheet.cell(row=i, column=3)
        s3.value = added_users[index].type

        s4 = sheet.cell(row=i, column=4)
        s4.value = added_users[index].name_usr

        last_idx = i + 2

    s_data = sheet.cell(row=last_idx, column=1)
    s_data.value = "Дата экспорта:"
    s_data = sheet.cell(row=last_idx, column=2)
    s_data.value = datetime.date.today()

    wb.save("files/Добавленные аккаунты.xlsx")
    doc = open("files/Добавленные аккаунты.xlsx", 'rb')

    await callback.answer()
    await callback.message.answer_document(doc)
